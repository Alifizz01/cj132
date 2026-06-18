"""Read per-cell condition layers from a sim_conditions Excel workbook.

One sheet per attribute (``layer_state``, ``layer_shade``, ``layer_life``,
``layer_incidence``), all the same shape as the master grid (n_rows × n_cols).
A blank cell = the schema default (healthy, 1.0, 1.0, 1.0).

Loader API
----------
``load_condition_layers(wb, *, n_rows, n_cols) -> dict[int, CellCondition]``

    Accepts either an open :class:`openpyxl.Workbook` or a ``Path``/``str``
    workbook path.  Returns ``{k: CellCondition}`` for every tile
    ``k = r * n_cols + c``.  Raises ``ValueError`` on shape mismatch.

The workbook template has one header row (row 1) and one label column (col A),
so data starts at Excel row 2, col B.  Row ``r`` (0-based) lives at Excel row
``r + 2``; column ``c`` (0-based) at Excel column ``c + 2``.

``generate_condition_workbook(path, *, n_rows, n_cols) -> None``

    Create a blank-but-structured sim_conditions.xlsx (all cells blank = default)
    with descriptive headers.  Only creates; never overwrites.

``normalize_condition_workbook(path, *, n_rows, n_cols) -> None``

    Open an existing workbook and add any missing layer sheets, preserving
    existing user values.  The converse of generate: idempotent in-place upgrade.
"""
from __future__ import annotations

import dataclasses
from pathlib import Path
from typing import Union

import openpyxl

from powerpy.simulation.cell_condition import CellCondition

# Canonical layer names (also the sheet names in the workbook).
LAYER_NAMES = ("layer_state", "layer_shade", "layer_life", "layer_incidence")

# Defaults as plain Python values (parallel to CellCondition defaults).
_DEFAULTS: dict[str, object] = {
    "layer_state":     "healthy",
    "layer_shade":     1.0,
    "layer_life":      1.0,
    "layer_incidence": 1.0,
}

# Map "active" alias to "healthy" for the state layer.
_STATE_ALIAS = {"active": "healthy", "healthy": "healthy", "failed_open": "failed_open"}


def load_condition_layers(
    wb: Union[openpyxl.Workbook, Path, str],
    *,
    n_rows: int,
    n_cols: int,
) -> dict[int, CellCondition]:
    """Build ``{k: CellCondition}`` from a sim_conditions workbook.

    Parameters
    ----------
    wb:
        An already-open :class:`openpyxl.Workbook` OR a filesystem path.
    n_rows, n_cols:
        The expected grid dimensions.  Every layer sheet must have exactly
        this shape (data rows/cols, ignoring the header row and label column).

    Returns
    -------
    dict mapping tile index ``k = r * n_cols + c`` to :class:`CellCondition`.
    All ``n_rows * n_cols`` tiles are present (blank → default).

    Raises
    ------
    ValueError
        If any layer sheet has a shape mismatch vs ``(n_rows, n_cols)``.
    """
    if not isinstance(wb, openpyxl.Workbook):
        wb = openpyxl.load_workbook(str(wb), data_only=True)

    # Read raw values from each layer sheet into a list-of-lists.
    raw: dict[str, list[list]] = {}
    for layer in LAYER_NAMES:
        if layer not in wb.sheetnames:
            # Missing sheet → treat entire layer as blank (all defaults).
            raw[layer] = [[None] * n_cols for _ in range(n_rows)]
            continue

        ws = wb[layer]
        # Validate shape BEFORE reading (ws.cell() materialises cells and
        # increases max_row/max_column, which would corrupt the check).
        # max_row / max_column give the last row/col that was ever written;
        # our template always writes the label column and header row, so
        # max_row-1 = data rows, max_col-1 = data cols.
        sheet_data_rows = ws.max_row - 1
        sheet_data_cols = ws.max_column - 1
        if sheet_data_rows != n_rows or sheet_data_cols != n_cols:
            raise ValueError(
                "shape mismatch in sheet %r: declared grid (%d, %d) but "
                "sheet has %d data rows and %d data cols"
                % (layer, n_rows, n_cols, sheet_data_rows, sheet_data_cols)
            )

        # Data starts at Excel row 2 (row 1 is header), col 2 (col 1 is label).
        data_rows = []
        for r in range(n_rows):
            row_vals = []
            for c in range(n_cols):
                cell = ws.cell(row=r + 2, column=c + 2)
                row_vals.append(cell.value)
            data_rows.append(row_vals)

        raw[layer] = data_rows

    # Build CellCondition map.
    conditions: dict[int, CellCondition] = {}
    default = CellCondition()

    for r in range(n_rows):
        for c in range(n_cols):
            k = r * n_cols + c

            raw_state = raw["layer_state"][r][c]
            raw_shade = raw["layer_shade"][r][c]
            raw_life = raw["layer_life"][r][c]
            raw_inc = raw["layer_incidence"][r][c]

            state_str = _parse_state(raw_state)
            shade = _parse_float(raw_shade, default.shade)
            life = _parse_float(raw_life, default.life)
            incidence = _parse_float(raw_inc, default.incidence)

            if (state_str == default.state and shade == default.shade
                    and life == default.life and incidence == default.incidence):
                conditions[k] = default
            else:
                conditions[k] = CellCondition(
                    state=state_str, shade=shade, life=life, incidence=incidence)

    return conditions


def _parse_state(raw) -> str:
    if raw is None:
        return "healthy"
    s = str(raw).strip().lower()
    if s == "" or s == "none":
        return "healthy"
    if s not in _STATE_ALIAS:
        raise ValueError(
            "layer_state: unrecognised state value %r (valid: %s)"
            % (raw, sorted(_STATE_ALIAS.keys())))
    return _STATE_ALIAS[s]


def _parse_float(raw, default: float) -> float:
    if raw is None:
        return default
    try:
        v = float(raw)
    except (TypeError, ValueError):
        raise ValueError("condition layer: expected a float, got %r" % (raw,))
    return v


# ---------------------------------------------------------------------------
# Workbook generation / normalization
# ---------------------------------------------------------------------------

_LAYER_DESCRIPTIONS = {
    "layer_state":     "Cell state: 'healthy' (default) or 'failed_open'. Blank = healthy.",
    "layer_shade":     "Irradiance transmittance [0..1]. 1.0 = unshaded (default). Blank = 1.0.",
    "layer_life":      "Mechanical chipping current-derate [0..1]. 1.0 = full (default). Blank = 1.0.",
    "layer_incidence": "Cosine factor for blanket-bow incidence [0..1]. 1.0 = normal (default). Blank = 1.0.",
}


def generate_condition_workbook(path: Union[Path, str], *, n_rows: int, n_cols: int) -> None:
    """Create a blank sim_conditions workbook at ``path``.

    Writes one sheet per layer with a descriptive header row and a label
    column.  All data cells are left blank (= the schema default for each
    attribute).  Raises ``FileExistsError`` if the file already exists.
    """
    path = Path(path)
    if path.exists():
        raise FileExistsError("workbook already exists: %s" % path)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for layer in LAYER_NAMES:
        ws = wb.create_sheet(layer)
        desc = _LAYER_DESCRIPTIONS.get(layer, "")
        # Row 1: description in A1, then column indices B1..
        ws.cell(row=1, column=1, value=desc)
        for c in range(n_cols):
            ws.cell(row=1, column=c + 2, value=c)
        # Rows 2..(n_rows+1): row label in col A, blank data in B..
        for r in range(n_rows):
            ws.cell(row=r + 2, column=1, value=r)
            # data cells left blank

    wb.save(str(path))


def normalize_condition_workbook(path: Union[Path, str], *, n_rows: int, n_cols: int) -> None:
    """Add any missing layer sheets to an existing workbook (in place).

    Existing sheets and their user values are preserved.  New sheets are
    added with a descriptive header; existing sheets are not resized.
    """
    path = Path(path)
    wb = openpyxl.load_workbook(str(path))
    changed = False

    for layer in LAYER_NAMES:
        if layer in wb.sheetnames:
            continue
        ws = wb.create_sheet(layer)
        desc = _LAYER_DESCRIPTIONS.get(layer, "")
        ws.cell(row=1, column=1, value=desc)
        for c in range(n_cols):
            ws.cell(row=1, column=c + 2, value=c)
        for r in range(n_rows):
            ws.cell(row=r + 2, column=1, value=r)
        changed = True

    if changed:
        wb.save(str(path))
