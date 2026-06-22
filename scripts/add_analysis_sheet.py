# -*- coding: utf-8 -*-
"""Add the `analysis` scope sheet and the string-shunt-diode reference to an
existing params.xlsx, in place (idempotent).

Mirrors the authoritative workbook screenshots from the 03_PowerPy_02 repo:

* a new `analysis` sheet -- the SCOPE: which (launch, phase) configs to actually
  investigate and include in the report, with their operating conditions;
* in `cell_params`: rename `diode_reference_file` -> `cell_shunt_diode_reference_file`
  and add `string_shunt_diode_reference_file` -> diodes/aRoche.json.

Usage:
    python scripts/add_analysis_sheet.py [path/to/params.xlsx]
"""
import sys
from pathlib import Path

import openpyxl

# analysis-sheet header + the two rows from the screenshot
_HEADER = ["launch", "phase", "season", "temperature",
           "string_loss", "sun_angle", "v_operating"]
_ROWS = [
    ["single", "End_of_Life", "1322/1367", 51.1, 1, 0, 101.5],
    ["dual",   "End_of_Life", "1322/1367", 51.1, 1, 0, 101.5],
]


def _add_analysis(wb) -> str:
    if "analysis" in wb.sheetnames:
        return "analysis: already present (left unchanged)"
    ws = wb.create_sheet("analysis")
    ws.append(_HEADER)
    for r in _ROWS:
        ws.append(r)
    return "analysis: created with %d config rows" % len(_ROWS)


def _patch_cell_params(wb) -> str:
    if "cell_params" not in wb.sheetnames:
        return "cell_params: NOT FOUND (skipped)"
    ws = wb["cell_params"]
    # find the param column (first column holds the key)
    params = {}
    for row in ws.iter_rows(min_row=2):
        key = row[0].value
        if key:
            params[str(key).strip()] = row
    msgs = []

    # rename diode_reference_file -> cell_shunt_diode_reference_file
    if "diode_reference_file" in params and "cell_shunt_diode_reference_file" not in params:
        cell = params["diode_reference_file"][0]
        cell.value = "cell_shunt_diode_reference_file"
        # name column (col B), if present
        if len(params["diode_reference_file"]) > 1:
            params["diode_reference_file"][1].value = "Cell Shunt Diode's reference File"
        msgs.append("renamed diode_reference_file -> cell_shunt_diode_reference_file")

    # add string_shunt_diode_reference_file row if absent
    have_string = any(k == "string_shunt_diode_reference_file" for k in params)
    if not have_string:
        # match the column shape of an existing path row (param,name,value,unit,type,source...)
        ncols = ws.max_column
        new = [None] * ncols
        new[0] = "string_shunt_diode_reference_file"
        if ncols > 1:
            new[1] = "String Shunt Diode's reference File"
        if ncols > 2:
            new[2] = "diodes/aRoche.json"
        # try to set a 'type' = path column if the sheet has one (look at header)
        header = [c.value for c in ws[1]]
        if "type" in header:
            new[header.index("type")] = "path"
        ws.append(new)
        msgs.append("added string_shunt_diode_reference_file -> diodes/aRoche.json")
    else:
        msgs.append("string_shunt_diode_reference_file already present")
    return "cell_params: " + "; ".join(msgs)


def main() -> int:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else None
    if path is None:
        root = Path(__file__).resolve().parent.parent
        for cand in (root / "src" / "powerpy" / "param" / "params.xlsx", root / "params.xlsx", root / "examples" / "params.xlsx"):
            if cand.is_file():
                path = cand
                break
    if path is None or not path.is_file():
        sys.exit("ERROR: params.xlsx not found; pass its path explicitly.")

    wb = openpyxl.load_workbook(path)
    print(_add_analysis(wb))
    print(_patch_cell_params(wb))
    wb.save(path)
    print("saved -> %s" % path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
