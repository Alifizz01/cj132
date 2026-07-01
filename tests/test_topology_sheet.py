"""The 'topology' sheet: uniform keys OR a grid_file reference, 'panel' fallback."""
from pathlib import Path

import openpyxl
import pytest

from powerpy.loader.sim_config import read_panel_config, read_topology, resolve_layout

ROOT = Path(__file__).resolve().parents[1]
PARAMS = ROOT / "src" / "powerpy" / "param" / "params.xlsx"
GRID = ROOT / "src" / "powerpy" / "data" / "layouts" / "grid_3x2x12.json"
pytestmark = pytest.mark.skipif(not PARAMS.exists(), reason="params.xlsx not present")


def _make_wb(path, sheet, rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(sheet)
    ws.append(["param", "value"])
    for r in rows:
        ws.append(list(r))
    wb.save(str(path))
    return path


def test_legacy_panel_sheet_fallback(tmp_path):
    p = _make_wb(tmp_path / "d.xlsx", "panel",
                 [("n_blocks", 2), ("n_parallel", 3), ("n_series", 5)])
    cfg = read_topology(p)
    assert (cfg["n_blocks"], cfg["n_parallel"], cfg["n_series"]) == (2, 3, 5)
    assert cfg["grid_file"] is None
    assert cfg == {**read_panel_config(p), "grid_file": None}


def test_topology_sheet_uniform_form(tmp_path):
    p = _make_wb(tmp_path / "d.xlsx", "topology",
                 [("n_blocks", 2), ("n_parallel", 3), ("n_series", 5),
                  ("imp_sigma", 0.01)])
    cfg = read_topology(p)
    assert (cfg["n_blocks"], cfg["n_parallel"], cfg["n_series"]) == (2, 3, 5)
    assert cfg["imp_sigma"] == 0.01
    assert cfg["grid_file"] is None
    lay = resolve_layout(cfg, base_dir=tmp_path)
    assert lay.n_tiles == 2 * 3 * 5


def test_topology_sheet_wins_over_panel(tmp_path):
    p = tmp_path / "d.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet, n in (("panel", 1), ("topology", 4)):
        ws = wb.create_sheet(sheet)
        ws.append(["param", "value"])
        ws.append(["n_parallel", n])
        ws.append(["n_series", 2])
    wb.save(str(p))
    assert read_topology(p)["n_parallel"] == 4


def test_topology_grid_file_form(tmp_path):
    p = _make_wb(tmp_path / "d.xlsx", "topology",
                 [("grid_file", str(GRID))])
    cfg = read_topology(p)
    assert cfg["grid_file"] == str(GRID)
    lay = resolve_layout(cfg, base_dir=tmp_path)
    assert lay.n_tiles == 72          # grid_3x2x12: 3 blocks x 2 parallel x 12 series


def test_grid_file_ignored_on_legacy_panel_sheet(tmp_path):
    """Pre-P2, read_panel_config ignored unknown keys -- so a stray grid_file
    row on a legacy 'panel' sheet must stay inert. grid_file is only honoured
    on the new 'topology' sheet."""
    p = _make_wb(tmp_path / "d.xlsx", "panel",
                 [("n_parallel", 4), ("n_series", 10),
                  ("grid_file", "leftover_experiment.json")])
    cfg = read_topology(p)
    assert cfg["grid_file"] is None
    lay = resolve_layout(cfg, base_dir=tmp_path)   # must NOT try to load the JSON
    assert lay.n_tiles == 40


def test_whitespace_grid_file_is_none(tmp_path):
    p = _make_wb(tmp_path / "d.xlsx", "topology",
                 [("n_parallel", 2), ("n_series", 3), ("grid_file", "   ")])
    assert read_topology(p)["grid_file"] is None


def test_grid_file_relative_to_workbook(tmp_path):
    import shutil
    shutil.copy(GRID, tmp_path / "mygrid.json")
    p = _make_wb(tmp_path / "d.xlsx", "topology", [("grid_file", "mygrid.json")])
    lay = resolve_layout(read_topology(p), base_dir=tmp_path)
    assert lay.n_tiles == 72
