# -*- coding: utf-8 -*-
"""Set specific values in the params.xlsx `mission_orbit` sheet (idempotent).

Updates existing params in place and appends any that are missing, preserving the
key-value schema  ``param | name | value | unit | type | source``.

Edit ``_CHANGES`` to change what it writes. Each entry is
``param: (value, name, unit, type)`` -- name/unit/type are used only when the row
has to be created.

Usage:
    python scripts/set_mission_orbit.py [path/to/params.xlsx]
"""
import sys
from pathlib import Path

import openpyxl

# param -> (value, name, unit, type)   (name/unit/type only used when CREATING)
_CHANGES = {
    "bus_voltage":        (101.5, "Bus Voltage", "V", "float"),
    "max_beta_angle_deg": (23.5,  "Max Beta Angle", "deg", "float"),
    "bond_albedo":        (0.30,  "Bond Albedo", "-", "float"),
    "planet_temp_k":      (255.0, "Planet Temperature", "K", "float"),
    "ir_emissivity":      (1.0,   "Planet IR Emissivity", "-", "float"),
}

_SHEET = "mission_orbit"


def main() -> int:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else None
    if path is None:
        root = Path(__file__).resolve().parent.parent
        for cand in (root / "params.xlsx", root / "examples" / "params.xlsx"):
            if cand.is_file():
                path = cand
                break
    if path is None or not path.is_file():
        sys.exit("ERROR: params.xlsx not found; pass its path explicitly.")

    wb = openpyxl.load_workbook(path)
    if _SHEET not in wb.sheetnames:
        sys.exit("ERROR: sheet '%s' not found in %s" % (_SHEET, path))
    ws = wb[_SHEET]

    header = [c.value for c in ws[1]]
    col = {name: i for i, name in enumerate(header)}  # 'param'->0, 'value'->2, ...
    if "param" not in col or "value" not in col:
        sys.exit("ERROR: '%s' is not a key-value sheet (need param/value columns)" % _SHEET)

    # index existing rows by param
    rows = {}
    for r in ws.iter_rows(min_row=2):
        key = r[col["param"]].value
        if key:
            rows[str(key).strip()] = r

    for param, (value, name, unit, type_) in _CHANGES.items():
        if param in rows:
            old = rows[param][col["value"]].value
            rows[param][col["value"]].value = value
            print("updated %-20s %s -> %s" % (param, old, value))
        else:
            new = [None] * ws.max_column
            new[col["param"]] = param
            new[col["value"]] = value
            if "name" in col:
                new[col["name"]] = name
            if "unit" in col:
                new[col["unit"]] = unit
            if "type" in col:
                new[col["type"]] = type_
            ws.append(new)
            print("added   %-20s = %s %s" % (param, value, unit))

    wb.save(path)
    print("saved -> %s" % path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
