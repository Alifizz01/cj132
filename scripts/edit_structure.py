# -*- coding: utf-8 -*-
"""Edit the params.xlsx `structure` sheet -- the report's table of contents.

The `structure` sheet is long-format, one row per report section::

    include | id | title | description | type | ref | audience

This tool edits it by ``id`` (idempotent). With no arguments it just LISTS the
current structure, so running it never changes anything by accident.

Examples:
    python scripts/edit_structure.py                       # list sections
    python scripts/edit_structure.py --off fig_iv_panel    # exclude a section
    python scripts/edit_structure.py --on  fig_iv_panel    # include it again
    python scripts/edit_structure.py --title results="EOL Results"
    python scripts/edit_structure.py --audience cell_params=engineer

To ADD new section rows, edit the ``_ADD`` list below and run with --apply-add.
Each entry's keys must match the sheet header (include/id/title/type/ref/audience).
"""
import argparse
import sys
from pathlib import Path

import openpyxl

_SHEET = "structure"

# New rows to append when run with --apply-add. Empty by default (no-op).
_ADD = [
    # {"include": True, "id": "requirement_table", "title": "Mission Requirements",
    #  "description": "", "type": "requirement_table", "ref": "", "audience": "both"},
]


def _find_params(argpath):
    if argpath:
        p = Path(argpath).expanduser().resolve()
        if not p.is_file():
            sys.exit("ERROR: params file not found: %s" % p)
        return p
    root = Path(__file__).resolve().parent.parent
    for cand in (root / "src" / "powerpy" / "param" / "params.xlsx", root / "params.xlsx", root / "examples" / "params.xlsx"):
        if cand.is_file():
            return cand
    sys.exit("ERROR: params.xlsx not found; pass --file.")


def _index(ws, col):
    rows = {}
    for r in ws.iter_rows(min_row=2):
        key = r[col["id"]].value
        if key:
            rows[str(key).strip()] = r
    return rows


def _kv(pairs):
    """['a=b', 'c=d'] -> {'a':'b','c':'d'} (value may contain '=')."""
    out = {}
    for p in pairs or []:
        if "=" not in p:
            sys.exit("ERROR: expected id=value, got %r" % p)
        k, v = p.split("=", 1)
        out[k.strip()] = v.strip()
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", default=None)
    ap.add_argument("--on", nargs="*", default=[], help="ids to include (include=TRUE)")
    ap.add_argument("--off", nargs="*", default=[], help="ids to exclude (include=FALSE)")
    ap.add_argument("--title", nargs="*", default=[], help="id=NewTitle ...")
    ap.add_argument("--audience", nargs="*", default=[], help="id=audience ...")
    ap.add_argument("--apply-add", action="store_true", help="append the _ADD rows")
    args = ap.parse_args()

    path = _find_params(args.file)
    wb = openpyxl.load_workbook(path)
    if _SHEET not in wb.sheetnames:
        sys.exit("ERROR: sheet '%s' not found in %s" % (_SHEET, path))
    ws = wb[_SHEET]
    header = [c.value for c in ws[1]]
    col = {name: i for i, name in enumerate(header)}
    for need in ("id", "include"):
        if need not in col:
            sys.exit("ERROR: '%s' sheet missing '%s' column" % (_SHEET, need))
    rows = _index(ws, col)

    changed = False

    def _require(ids):
        missing = [i for i in ids if i not in rows]
        if missing:
            sys.exit("ERROR: unknown section id(s): %s\nknown: %s"
                     % (", ".join(missing), ", ".join(rows)))

    _require(list(args.on) + list(args.off)
             + list(_kv(args.title)) + list(_kv(args.audience)))

    for i in args.on:
        rows[i][col["include"]].value = True
        print("include %s = TRUE" % i); changed = True
    for i in args.off:
        rows[i][col["include"]].value = False
        print("include %s = FALSE" % i); changed = True
    for i, t in _kv(args.title).items():
        rows[i][col["title"]].value = t
        print("title   %s = %r" % (i, t)); changed = True
    for i, a in _kv(args.audience).items():
        rows[i][col["audience"]].value = a
        print("audience %s = %r" % (i, a)); changed = True

    if args.apply_add:
        for spec in _ADD:
            if spec.get("id") in rows:
                print("skip add %s (already present)" % spec.get("id"))
                continue
            new = [None] * ws.max_column
            for k, v in spec.items():
                if k in col:
                    new[col[k]] = v
            ws.append(new)
            print("added section %s" % spec.get("id")); changed = True
        if not _ADD:
            print("--apply-add: _ADD list is empty (nothing to add)")

    if changed:
        try:
            wb.save(path)
        except PermissionError:
            sys.exit("ERROR: could not write %s -- it is open in Excel (or locked). "
                     "Close it and re-run; no changes were saved." % path.name)
        print("saved -> %s" % path)
    else:
        # list-only mode
        print("structure of %s:" % path.name)
        for key, r in rows.items():
            inc = r[col["include"]].value
            typ = r[col["type"]].value if "type" in col else ""
            ref = r[col["ref"]].value if "ref" in col else ""
            mark = "[x]" if inc in (True, 1, "TRUE", "true") else "[ ]"
            print("  %s %-16s type=%-14s ref=%s" % (mark, key, typ, ref or ""))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
