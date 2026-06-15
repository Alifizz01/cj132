# -*- coding: utf-8 -*-
"""Generator: read the current params.xlsx and emit a standalone build_params.py
that recreates every sheet exactly. Run once; the output is the deliverable.
"""
import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "params.xlsx"
OUT = ROOT / "scripts" / "build_params.py"

wb = openpyxl.load_workbook(SRC, data_only=True)

lines = []
lines.append("# -*- coding: utf-8 -*-")
lines.append('"""Build the COMPLETE params.xlsx (every sheet) from scratch.')
lines.append("")
lines.append("Auto-generated from the working params.xlsx, so it reproduces all sheets")
lines.append("exactly: document, cell_params, mission_orbit, mission_param, sections,")
lines.append("losses, radiation_fluxes, structure, analysis, requirement.")
lines.append("")
lines.append("Edit the SHEETS data below to change values, then re-run to rebuild.")
lines.append("")
lines.append("Usage:")
lines.append("    python scripts/build_params.py                 # -> params.xlsx (repo root)")
lines.append("    python scripts/build_params.py out.xlsx        # -> a chosen path")
lines.append('"""')
lines.append("import datetime  # noqa: F401  (used by baked-in date values)")
lines.append("import sys")
lines.append("from pathlib import Path")
lines.append("")
lines.append("import openpyxl")
lines.append("")
lines.append("# Each entry: (sheet_name, [row, row, ...]); row[0] is the header row.")
lines.append("SHEETS = [")
for name in wb.sheetnames:
    ws = wb[name]
    lines.append("    (%r, [" % name)
    for row in ws.iter_rows(values_only=True):
        # trim wholly-empty trailing rows
        if row is None or all(c is None for c in row):
            continue
        lines.append("        %r," % (list(row),))
    lines.append("    ]),")
lines.append("]")
lines.append("")
lines.append("")
lines.append("def main() -> int:")
lines.append("    out = Path(sys.argv[1]) if len(sys.argv) > 1 else \\")
lines.append("        Path(__file__).resolve().parent.parent / 'params.xlsx'")
lines.append("    wb = openpyxl.Workbook()")
lines.append("    wb.remove(wb.active)")
lines.append("    for name, rows in SHEETS:")
lines.append("        ws = wb.create_sheet(name)")
lines.append("        for r in rows:")
lines.append("            ws.append(list(r))")
lines.append("    try:")
lines.append("        wb.save(out)")
lines.append("    except PermissionError:")
lines.append("        sys.exit('ERROR: %s is open/locked in Excel. Close it or pass another path.' % out)")
lines.append("    print('built %d sheets -> %s' % (len(SHEETS), out))")
lines.append("    return 0")
lines.append("")
lines.append("")
lines.append("if __name__ == '__main__':")
lines.append("    raise SystemExit(main())")
lines.append("")

OUT.write_text("\n".join(lines), encoding="utf-8")
print("wrote %s (%d sheets)" % (OUT, len(wb.sheetnames)))
