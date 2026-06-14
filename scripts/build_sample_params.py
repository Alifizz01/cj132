"""Build a sample params.xlsx with all required sheets, for testing the loader."""
from datetime import date
from pathlib import Path

from openpyxl import Workbook


def build_sample_workbook(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    _build_document_meta(wb)
    _build_revision_history(wb)
    _build_cell_params(wb)
    _build_mission_params(wb)
    _build_sections(wb)
    _build_array_topology(wb)
    _build_panels(wb)
    _build_losses(wb)
    _build_radiation_fluxes(wb)

    wb.save(path)


def _build_document_meta(wb):
    ws = wb.create_sheet("document_meta")
    ws.append(["param", "name", "value", "type", "notes"])
    rows = [
        ("doc_number",      "Document Number",  "G2G-ADSO-AN-10003",                       "string", ""),
        ("doc_title",       "Document Title",   "Solar Array Analysis - Mission XYZ",      "string", ""),
        ("issue",           "Issue",             2,                                          "int",    ""),
        ("issue_date",      "Issue Date",       date(2026, 5, 16),                          "date",   ""),
        ("project_name",    "Project Name",     "XYZ Telecom Satellite",                    "string", ""),
        ("project_code",    "Project Code",     "XYZ-001",                                  "string", ""),
        ("customer",        "Customer",         "ESA",                                      "string", ""),
        ("contract_ref",    "Contract Reference","4000123456/22/NL/AB",                    "string", ""),
        ("classification",  "Classification",   "Airbus Proprietary",                       "string", ""),
        ("export_control",  "Export Control",   "Not subject to ITAR",                      "string", ""),
        ("prepared_by",     "Prepared By",      "M. A. I. bin Ibrahim",                     "string", ""),
        ("prepared_role",   "Preparer Role",    "Werkstudent, Solar Arrays Engineering",    "string", ""),
        ("checked_by",      "Checked By",       "",                                         "string", ""),
        ("checked_role",    "Checker Role",     "",                                         "string", ""),
        ("approved_by",     "Approved By",      "",                                         "string", ""),
        ("approved_role",   "Approver Role",    "",                                         "string", ""),
        ("logo_file",       "Logo File",        "assets/airbus_logo.png",                   "path",   ""),
        ("template_style",  "Template Style",   "airbus_default",                           "string", ""),
    ]
    for r in rows:
        ws.append(list(r))


def _build_revision_history(wb):
    ws = wb.create_sheet("revision_history")
    ws.append(["issue", "date", "author", "description", "status"])
    ws.append([1, date(2026, 3, 1),  "M. A. I. bin Ibrahim", "Initial issue.",                                                            "superseded"])
    ws.append([2, date(2026, 5, 16), "M. A. I. bin Ibrahim", "Updated radiation fluxes per ECSS-E-ST-10-04C Rev.1; added EOL margin.",   "current"])


def _build_cell_params(wb):
    ws = wb.create_sheet("cell_params")
    ws.append(["param", "name", "value", "unit", "type", "source"])
    rows = [
        ("cell_name",            "Cell Name",            "3G30LARS GEO",                  "-",   "string", "datasheet"),
        ("cell_reference_file",  "Cell Reference File",  "cells/3G30LARS_GEO.json",       "-",   "path",   "-"),
        ("diode_reference_file", "Diode Reference File", "diodes/external.json",          "-",   "path",   "-"),
        ("manufacturer",         "Manufacturer",         "Azur Space Solar Power",        "-",   "string", "datasheet"),
        ("base_material",        "Base Material",        "GaInP2/GaAs on Ge Substrate",   "-",   "string", "datasheet"),
        ("junction",             "Junction",             "DEMO",                          "-",   "string", "-"),
        ("ar_coating",           "A/R Coating",          "TiOx / Al2O3",                  "-",   "string", "datasheet"),
        ("front_contact",        "Front Contact",        "Au / Ag / AuZn weldable",       "-",   "string", "datasheet"),
        ("rear_contact",         "Rear Contact",         "AuGe / Ag / Au weldable",       "-",   "string", "datasheet"),
        ("substrate_material",   "Substrate Material",   "p-Ge",                          "-",   "string", "datasheet"),
        ("substrate_thickness",  "Substrate Thickness",  140,                             "um",  "float",  "datasheet"),
        ("cell_length",          "Cell Length",          124.3,                           "mm",  "float",  "datasheet"),
        ("cell_width",           "Cell Width",           60.5,                            "mm",  "float",  "datasheet"),
        ("cell_thickness",       "Cell Thickness",       160,                             "um",  "float",  "datasheet"),
        ("cell_area",            "Cell Area",            70.75,                           "cm2", "float",  "datasheet"),
        ("cell_mass",            "Cell Mass",            6100,                            "mg",  "float",  "datasheet"),
    ]
    for r in rows:
        ws.append(list(r))


def _build_mission_params(wb):
    ws = wb.create_sheet("mission_params")
    ws.append(["param", "name", "value", "unit", "type", "source"])
    rows = [
        ("orbit_type",            "Orbit Type",                "GEO",                 "-",     "string", ""),
        ("altitude_km",           "Altitude",                  35786,                 "km",    "float",  ""),
        ("inclination_deg",       "Inclination",               0.0,                   "deg",   "float",  ""),
        ("eclipse_duration_min",  "Max Eclipse Duration",      72,                    "min",   "float",  ""),
        ("launch_date",           "Launch Date",               date(2027, 3, 15),     "-",     "date",   ""),
        ("mission_duration_years","Mission Duration",          15,                    "years", "float",  ""),
        ("leop_duration_days",    "LEOP Duration",             10,                    "days",  "float",  ""),
        ("orp_duration_months",   "ORP Duration",              3,                     "months","float",  ""),
        ("sun_intensity_bol",     "Sun Intensity BOL",         1367,                  "W/m2",  "float",  ""),
        ("sun_intensity_eol_min", "Sun Intensity EOL Min",     1322,                  "W/m2",  "float",  ""),
        ("temp_max_K",            "Max Temperature",           363,                   "K",     "float",  ""),
        ("temp_min_K",            "Min Temperature",           173,                   "K",     "float",  ""),
        ("bus_voltage",           "Bus Voltage",               100,                   "V",     "float",  ""),
        ("max_string_current",    "Max String Current",        2.5,                   "A",     "float",  ""),
    ]
    for r in rows:
        ws.append(list(r))


def _build_sections(wb):
    ws = wb.create_sheet("sections")
    ws.append([
        "section_id", "section_name",
        "n_strings_parallel", "n_sca_series_per_string", "n_sca_per_section",
        "include", "notes",
    ])
    ws.append(["section_a", "Section A", 4, 54, "=C2*D2", True, ""])
    ws.append(["section_b", "Section B", 3, 54, "=C3*D3", True, ""])
    ws.append(["section_c", "Section C", 3, 54, "=C4*D4", True, ""])
    ws.append(["section_d", "Section D", 3, 54, "=C5*D5", True, ""])


def _build_array_topology(wb):
    ws = wb.create_sheet("array_topology")
    ws.append(["param", "value", "unit", "type", "description"])
    ws.append(["n_wings", 2, "-", "int", "Number of wings"])
    ws.append(["n_panels_per_wing", 3, "-", "int", "Panels per wing"])


def _build_panels(wb):
    ws = wb.create_sheet("panels")
    ws.append([
        "panel_instance_id", "override_section_id", "action",
        "n_strings_parallel", "n_sca_series_per_string", "notes",
    ])
    # intentionally empty body — symmetric case


def _build_losses(wb):
    ws = wb.create_sheet("losses")
    ws.append([
        "name", "phase", "value", "level", "unit",
        "description", "source", "include",
    ])

    phases = ["BOL_ATC", "BOL_bc", "End_of_LEOP", "End_of_ORP", "End_of_Life"]

    # Coverglass loss: constant across phases, cell level
    for ph in phases:
        ws.append(["Coverglass loss", ph, 0.988, "cell", "-",
                   "Cover glass transmission", "AZUR datasheet", True])

    # RSS: varies a bit between phases
    rss_values = {"BOL_ATC": 0.978, "BOL_bc": 0.995,
                  "End_of_LEOP": 0.978, "End_of_ORP": 0.978, "End_of_Life": 0.978}
    for ph, v in rss_values.items():
        ws.append(["RSS", ph, v, "cell", "-",
                   "Radiation-shielded shunt diode leakage", "Internal estimate", True])

    # UV & micrometeorite: cell level
    for ph in phases:
        ws.append(["UV & Micrometeorite", ph, 0.975, "cell", "-",
                   "UV darkening + micrometeorite", "Internal estimate", True])

    # String loss: only at End_of_Life, string level
    ws.append(["String loss", "End_of_Life", 0.987, "string", "-",
               "Soldering interconnect series loss", "PS-PV-004", True])

    # SADM misalignment: array level, all phases
    for ph in phases:
        ws.append(["SADM misalignment", ph, 1.000, "array", "-",
                   "Solar array drive mechanism misalignment", "ECSS-ECSSI", True])


def _build_radiation_fluxes(wb):
    ws = wb.create_sheet("radiation_fluxes")
    ws.append([
        "name", "launch_config", "phase", "param", "value", "unit",
        "description", "source", "include",
    ])

    # From Table 4 in the source document
    data = [
        # single launch
        ("fluence", "single", "End_of_LEOP", "isc", 1.91e12),
        ("fluence", "single", "End_of_LEOP", "voc", 2.45e12),
        ("fluence", "single", "End_of_ORP",  "isc", 1.10e13),
        ("fluence", "single", "End_of_ORP",  "voc", 1.19e13),
        ("fluence", "single", "End_of_Life", "isc", 1.13e15),
        ("fluence", "single", "End_of_Life", "voc", 1.36e15),
        # dual launch
        ("fluence", "dual",   "End_of_LEOP", "isc", 1.67e14),
        ("fluence", "dual",   "End_of_LEOP", "voc", 4.73e14),
        ("fluence", "dual",   "End_of_ORP",  "isc", 1.34e15),
        ("fluence", "dual",   "End_of_ORP",  "voc", 3.76e15),
        ("fluence", "dual",   "End_of_Life", "isc", 2.46e15),
        ("fluence", "dual",   "End_of_Life", "voc", 5.11e15),
    ]
    for name, lc, ph, param, val in data:
        ws.append([name, lc, ph, param, val, "e/cm2",
                   "1 MeV equivalent electron fluence", "Table 4", True])


if __name__ == "__main__":
    import sys
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("params.xlsx")
    build_sample_workbook(out)
    print(f"Wrote {out}")
