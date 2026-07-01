"""Use case B must run from design.xlsx alone (cell + topology) plus a
scenario.xlsx holding only condition layers -- with NO report sheets anywhere.
This is the coupling P2 exists to break: before, write_results called
load_report_data and crashed on a workbook without document/losses/etc."""
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]
PARAMS = ROOT / "src" / "powerpy" / "param" / "params.xlsx"
pytestmark = pytest.mark.skipif(not PARAMS.exists(), reason="params.xlsx not present")

sys.path.insert(0, str(ROOT / "scripts"))
import write_results  # noqa: E402


def _copy_sheets(src_path, dst_path, names):
    src = openpyxl.load_workbook(str(src_path), data_only=True)
    dst = openpyxl.Workbook()
    dst.remove(dst.active)
    for name in names:
        ws = dst.create_sheet(name)
        for row in src[name].iter_rows(values_only=True):
            ws.append(list(row))
    dst.save(str(dst_path))
    return dst_path


def test_write_results_runs_without_report_sheets(tmp_path):
    design = _copy_sheets(PARAMS, tmp_path / "design.xlsx", ["cell_params", "panel"])
    scenario = _copy_sheets(PARAMS, tmp_path / "scenario.xlsx",
                            ["layer_state", "layer_shade", "layer_life", "layer_incidence"])
    out = tmp_path / "results.xlsx"

    rc = write_results.main(["--design", str(design), "--scenario", str(scenario),
                             "--out", str(out), "--no-report"])
    assert rc == 0
    wb = openpyxl.load_workbook(str(out), read_only=True)
    assert set(wb.sheetnames) == {"summary", "strings", "cells"}
    summary = {r[0]: r[1] for r in wb["summary"].iter_rows(values_only=True)}
    assert float(summary["Pmpp_W"]) > 0.0


def test_write_results_no_longer_imports_mega_loader():
    src = (ROOT / "scripts" / "write_results.py").read_text(encoding="utf-8")
    assert "load_report_data" not in src


def test_setup_sim_creates_missing_conditions_workbook(tmp_path):
    """Legacy contract (setup_sim Responsibility #1): a not-yet-existing --wb is
    GENERATED, not rejected. P2's find_workbooks regression-guarded here."""
    import setup_sim
    fresh = tmp_path / "fresh_conditions.xlsx"
    rc = setup_sim.main(["--layout", str(ROOT / "src/powerpy/data/layouts/simple_3block.json"),
                         "--wb", str(fresh),
                         "--runs", str(tmp_path / "runs"), "--run-id", "t"])
    assert rc == 0
    assert fresh.exists()
    assert (tmp_path / "runs" / "t" / "snapshot.json").exists()


def test_setup_sim_creates_missing_scenario_workbook(tmp_path):
    """Split-mode equivalent: --scenario may not exist yet (it is the conditions
    file setup_sim itself generates); --design must exist."""
    import setup_sim
    design = _copy_sheets(PARAMS, tmp_path / "design.xlsx", ["cell_params", "panel"])
    fresh = tmp_path / "scenario.xlsx"
    rc = setup_sim.main(["--design", str(design), "--scenario", str(fresh),
                         "--runs", str(tmp_path / "runs"), "--run-id", "t"])
    assert rc == 0
    assert fresh.exists()
    assert (tmp_path / "runs" / "t" / "snapshot.json").exists()
