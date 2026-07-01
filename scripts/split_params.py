"""Split one params.xlsx into the two-file database: design.xlsx + scenario.xlsx.

design.xlsx  = the hardware  (cell_params, sections, panel -> also copied as 'topology')
scenario.xlsx = the run      (losses, fluxes, mission, analysis, requirement,
                              document, structure, layer_* condition sheets)

Values are copied verbatim (values-only). Unknown sheets go to scenario.xlsx
with a warning. Refuses to overwrite existing outputs unless --overwrite.

Run (from project root):
    python scripts/split_params.py [params.xlsx] [--out-dir DIR] [--overwrite]
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT / "src"))

import openpyxl

# routing (see docs/superpowers/plans/2026-07-01-p2-two-file-database.md)
DESIGN_SHEETS = ("cell_params", "sections", "panel")
SCENARIO_SHEETS = ("losses", "radiation_fluxes", "mission_orbit", "mission_param",
                   "analysis", "requirement", "document", "structure",
                   "layer_state", "layer_shade", "layer_life", "layer_incidence")


def _copy_sheet(src_ws, dst_wb, title):
    ws = dst_wb.create_sheet(title)
    for row in src_ws.iter_rows(values_only=True):
        ws.append(list(row))
    return ws


def split_params(params_path, design_path, scenario_path, *,
                 overwrite: bool = False) -> tuple[Path, Path]:
    """Split ``params_path`` into the (design, scenario) pair. Returns the paths."""
    params_path = Path(params_path)
    design_path, scenario_path = Path(design_path), Path(scenario_path)
    for out in (design_path, scenario_path):
        if out.exists() and not overwrite:
            raise FileExistsError("refusing to overwrite %s (pass overwrite=True)" % out)

    src = openpyxl.load_workbook(str(params_path), data_only=True)
    design = openpyxl.Workbook(); design.remove(design.active)
    scenario = openpyxl.Workbook(); scenario.remove(scenario.active)

    for name in src.sheetnames:
        if name in DESIGN_SHEETS:
            _copy_sheet(src[name], design, name)
        elif name in SCENARIO_SHEETS:
            _copy_sheet(src[name], scenario, name)
        else:
            print("split_params: unknown sheet %r -> scenario.xlsx" % name)
            _copy_sheet(src[name], scenario, name)

    # the new 'topology' sheet starts as a verbatim copy of 'panel' (its loader
    # falls back to 'panel' anyway; the copy makes the new name authoritative).
    if "panel" in src.sheetnames:
        _copy_sheet(src["panel"], design, "topology")

    design.save(str(design_path))
    scenario.save(str(scenario_path))
    return design_path, scenario_path


def main(argv=None) -> int:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("params", nargs="?",
                   default=str(_ROOT / "src" / "powerpy" / "param" / "params.xlsx"))
    p.add_argument("--out-dir", default=None,
                   help="output directory (default: alongside the input workbook)")
    p.add_argument("--overwrite", action="store_true")
    a = p.parse_args(argv)
    params = Path(a.params)
    out_dir = Path(a.out_dir) if a.out_dir else params.parent
    d, s = split_params(params, out_dir / "design.xlsx", out_dir / "scenario.xlsx",
                        overwrite=a.overwrite)
    print("wrote %s\nwrote %s" % (d, s))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
