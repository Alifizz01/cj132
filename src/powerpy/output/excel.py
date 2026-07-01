"""The results.xlsx writer -- use case B's Excel output.

One workbook, three sheets: ``summary`` (array operating point + loss vs
nominal), ``strings`` (per-string current / node voltage / power), ``cells``
(per-cell k / state / shade / life / V / I / P). Promoted from
scripts/write_results.py so every output format lives in powerpy.output.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl


def write_results_xlsx(path: str | Path, layout_name: str,
                       summary: dict, strings, cells) -> Path:
    """Write the summary/strings/cells workbook. Returns the path."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    s = wb.create_sheet("summary")
    s.append(["layout", layout_name])
    for kk, vv in summary.items():
        s.append([kk, vv])
    st = wb.create_sheet("strings")
    st.append(["string_id", "current_A", "node_V", "power_W"])
    for r in strings:
        st.append(list(r))
    cs = wb.create_sheet("cells")
    cs.append(["cell_k", "string_id", "state", "shade", "life", "V_cell", "I_cell", "P_cell_W"])
    for r in cells:
        cs.append(list(r))
    wb.save(str(path))
    return Path(path)
