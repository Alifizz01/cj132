"""Phase 3 tests: condition-layer reader + setup_sim snapshot emitter (TDD)."""
from __future__ import annotations

import json
import math
from pathlib import Path

import openpyxl
import pytest

from powerpy.config.layout import panel_from_topology
from powerpy.simulation.cell_condition import CellCondition
from powerpy.loader.condition_layers import (
    load_condition_layers,
    LAYER_NAMES,
)

ROOT = Path(__file__).resolve().parents[1]
PARAMS = ROOT / "params.xlsx"
DATA = ROOT / "src" / "powerpy" / "data"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_workbook(n_rows: int, n_cols: int, overrides: dict | None = None) -> openpyxl.Workbook:
    """Create a minimal sim_conditions.xlsx in-memory (no file I/O).

    ``overrides`` is a dict mapping (layer_name, r, c) -> cell_value (1-based),
    to write specific non-blank values.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name in LAYER_NAMES:
        ws = wb.create_sheet(name)
        # Write a header comment row with column indices (optional, row 1)
        # Row 2 onward are the grid values (row 1 is the header)
        ws.append(["r\\c"] + list(range(n_cols)))  # header row 1
        for r in range(n_rows):
            row = [r] + [None] * n_cols
            ws.append(row)
    if overrides:
        for (layer, r, c), val in overrides.items():
            ws = wb[layer]
            # data starts at Excel row 2 (header at row 1), col 2 (label at col 1)
            ws.cell(row=r + 2, column=c + 2, value=val)
    return wb


# ---------------------------------------------------------------------------
# Task 1: load_condition_layers
# ---------------------------------------------------------------------------

class TestLoadConditionLayers:
    def test_blank_sheet_gives_all_defaults(self):
        """A fully-blank workbook gives CellCondition() for every tile k."""
        n_rows, n_cols = 2, 3
        wb = _make_workbook(n_rows, n_cols)
        conds = load_condition_layers(wb, n_rows=n_rows, n_cols=n_cols)
        n_tiles = n_rows * n_cols
        assert set(conds.keys()) == set(range(n_tiles))
        default = CellCondition()
        for k, c in conds.items():
            assert c == default, f"tile {k}: expected default, got {c}"

    def test_shade_written_to_sheet_is_read_back(self):
        n_rows, n_cols = 2, 3
        wb = _make_workbook(n_rows, n_cols,
                             overrides={("layer_shade", 0, 1): 0.5})
        conds = load_condition_layers(wb, n_rows=n_rows, n_cols=n_cols)
        k_target = 0 * n_cols + 1   # row 0, col 1
        assert conds[k_target].shade == pytest.approx(0.5)
        # Other tiles still default
        assert conds[0].shade == pytest.approx(1.0)

    def test_life_written_to_sheet_is_read_back(self):
        n_rows, n_cols = 2, 3
        wb = _make_workbook(n_rows, n_cols,
                             overrides={("layer_life", 1, 2): 0.7})
        conds = load_condition_layers(wb, n_rows=n_rows, n_cols=n_cols)
        k_target = 1 * n_cols + 2
        assert conds[k_target].life == pytest.approx(0.7)

    def test_incidence_written_to_sheet_is_read_back(self):
        n_rows, n_cols = 2, 3
        wb = _make_workbook(n_rows, n_cols,
                             overrides={("layer_incidence", 0, 0): 0.866})
        conds = load_condition_layers(wb, n_rows=n_rows, n_cols=n_cols)
        assert conds[0].incidence == pytest.approx(0.866, abs=1e-6)

    def test_state_failed_open_written_and_read(self):
        n_rows, n_cols = 2, 3
        wb = _make_workbook(n_rows, n_cols,
                             overrides={("layer_state", 1, 0): "failed_open"})
        conds = load_condition_layers(wb, n_rows=n_rows, n_cols=n_cols)
        k_target = 1 * n_cols + 0
        assert conds[k_target].state == "failed_open"
        assert conds[0].state == "healthy"

    def test_shape_mismatch_raises(self):
        """If the workbook is the wrong size for the declared (n_rows, n_cols), raise."""
        n_rows, n_cols = 2, 3
        wb = _make_workbook(n_rows, n_cols)
        with pytest.raises(ValueError, match="shape mismatch"):
            load_condition_layers(wb, n_rows=3, n_cols=3)

    def test_multiple_attributes_on_same_tile(self):
        n_rows, n_cols = 3, 3
        wb = _make_workbook(n_rows, n_cols, overrides={
            ("layer_shade", 1, 1): 0.3,
            ("layer_life", 1, 1): 0.9,
            ("layer_incidence", 1, 1): 0.8,
        })
        conds = load_condition_layers(wb, n_rows=n_rows, n_cols=n_cols)
        k = 1 * n_cols + 1
        c = conds[k]
        assert c.shade == pytest.approx(0.3)
        assert c.life == pytest.approx(0.9)
        assert c.incidence == pytest.approx(0.8)

    def test_active_state_in_sheet_is_valid(self):
        """Explicit 'active' in state layer is treated as 'healthy'."""
        n_rows, n_cols = 2, 2
        wb = _make_workbook(n_rows, n_cols,
                             overrides={("layer_state", 0, 0): "active"})
        conds = load_condition_layers(wb, n_rows=n_rows, n_cols=n_cols)
        assert conds[0].state == "healthy"


# ---------------------------------------------------------------------------
# Task 2: setup_sim.py snapshot emitter
# ---------------------------------------------------------------------------

class TestSetupSimSnapshot:
    def test_snapshot_folder_created_with_json(self, tmp_path):
        """run_setup_sim() writes a resolved JSON into runs/<run_id>/."""
        from scripts.setup_sim import run_setup_sim

        layout = panel_from_topology(n_blocks=1, n_parallel=2, n_series=3)
        n_rows, n_cols = layout.n_rows, layout.n_cols
        wb = _make_workbook(n_rows, n_cols,
                             overrides={("layer_shade", 0, 1): 0.6})
        wb_path = tmp_path / "sim_conditions.xlsx"
        wb.save(str(wb_path))

        run_id = "test_run_1"
        run_setup_sim(layout=layout, wb_path=wb_path,
                      runs_dir=tmp_path / "runs", run_id=run_id)

        snap_dir = tmp_path / "runs" / run_id
        assert snap_dir.is_dir()
        snap_json = snap_dir / "snapshot.json"
        assert snap_json.is_file()

        data = json.loads(snap_json.read_text())
        assert "spec" in data
        assert "conditions" in data

    def test_snapshot_conditions_include_shaded_tile(self, tmp_path):
        from scripts.setup_sim import run_setup_sim

        layout = panel_from_topology(n_blocks=1, n_parallel=2, n_series=3)
        n_rows, n_cols = layout.n_rows, layout.n_cols
        wb = _make_workbook(n_rows, n_cols,
                             overrides={("layer_shade", 0, 1): 0.4})
        wb_path = tmp_path / "sim_conditions.xlsx"
        wb.save(str(wb_path))

        run_setup_sim(layout=layout, wb_path=wb_path,
                      runs_dir=tmp_path / "runs", run_id="r1")

        snap = json.loads((tmp_path / "runs" / "r1" / "snapshot.json").read_text())
        # tile at (row=0, col=1) -> k = 0*n_cols+1 = 1
        k_shaded = str(0 * n_cols + 1)
        assert k_shaded in snap["conditions"]
        assert snap["conditions"][k_shaded]["shade"] == pytest.approx(0.4)

    def test_snapshot_is_deterministic(self, tmp_path):
        from scripts.setup_sim import run_setup_sim

        layout = panel_from_topology(n_blocks=1, n_parallel=2, n_series=3)
        n_rows, n_cols = layout.n_rows, layout.n_cols
        wb = _make_workbook(n_rows, n_cols)
        wb_path = tmp_path / "sim_conditions.xlsx"
        wb.save(str(wb_path))

        run_setup_sim(layout=layout, wb_path=wb_path,
                      runs_dir=tmp_path / "runs", run_id="r1")
        run_setup_sim(layout=layout, wb_path=wb_path,
                      runs_dir=tmp_path / "runs2", run_id="r1")

        j1 = (tmp_path / "runs" / "r1" / "snapshot.json").read_text()
        j2 = (tmp_path / "runs2" / "r1" / "snapshot.json").read_text()
        assert j1 == j2

    def test_workbook_generate_when_missing(self, tmp_path):
        """If the workbook doesn't exist, setup_sim generates it."""
        from scripts.setup_sim import run_setup_sim

        layout = panel_from_topology(n_blocks=1, n_parallel=2, n_series=3)
        wb_path = tmp_path / "sim_conditions.xlsx"
        assert not wb_path.exists()

        run_setup_sim(layout=layout, wb_path=wb_path,
                      runs_dir=tmp_path / "runs", run_id="r1")

        assert wb_path.exists(), "setup_sim should create wb when missing"

    def test_workbook_not_clobbered_when_existing(self, tmp_path):
        """If the workbook already has data, existing values are preserved."""
        from scripts.setup_sim import run_setup_sim

        layout = panel_from_topology(n_blocks=1, n_parallel=2, n_series=3)
        n_rows, n_cols = layout.n_rows, layout.n_cols
        wb = _make_workbook(n_rows, n_cols,
                             overrides={("layer_shade", 0, 0): 0.5})
        wb_path = tmp_path / "sim_conditions.xlsx"
        wb.save(str(wb_path))

        run_setup_sim(layout=layout, wb_path=wb_path,
                      runs_dir=tmp_path / "runs", run_id="r1")

        wb2 = openpyxl.load_workbook(str(wb_path))
        ws = wb2["layer_shade"]
        # data cell at (row 0, col 0) -> Excel row=2, col=2
        val = ws.cell(row=2, column=2).value
        assert val == pytest.approx(0.5), "User data must be preserved"


# ---------------------------------------------------------------------------
# Task 3: round-trip integration (requires params.xlsx)
# ---------------------------------------------------------------------------

@pytest.mark.skipif(not PARAMS.exists(), reason="params.xlsx not present")
class TestRoundTrip:
    def test_shaded_tile_lowers_pmax(self, tmp_path):
        """layers -> conditions -> build_array_from_spec: shaded tile lowers Pmax."""
        from powerpy.loader.report import load_report_data
        from powerpy.simulation.spec_adapt import adapt_grid
        from powerpy.simulation.spec_build import build_array_from_spec
        from powerpy.simulation.environment import Environment

        cell = load_report_data(PARAMS, DATA).cell
        layout = panel_from_topology(n_blocks=1, n_parallel=2, n_series=3)
        n_rows, n_cols = layout.n_rows, layout.n_cols
        spec = adapt_grid(layout)
        env = Environment(temperature_c=28.0)

        def _pmax(conds):
            arr = build_array_from_spec(cell, spec, conditions=conds)
            arr.apply(env)
            v, i = arr.iv_curve()
            return float((v * i).max())

        # baseline: all blank (default conditions)
        nominal = _pmax({})

        # shaded: tile k=1 shaded to 40%
        wb = _make_workbook(n_rows, n_cols,
                             overrides={("layer_shade", 0, 1): 0.4})
        conds = load_condition_layers(wb, n_rows=n_rows, n_cols=n_cols)
        shaded_pmax = _pmax(conds)

        assert shaded_pmax < nominal, (
            "shaded tile should lower Pmax: nominal=%g shaded=%g" % (nominal, shaded_pmax))
