import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]
PARAMS = ROOT / "src" / "powerpy" / "param" / "params.xlsx"
pytestmark = pytest.mark.skipif(not PARAMS.exists(), reason="params.xlsx not present")

sys.path.insert(0, str(ROOT / "scripts"))
from split_params import split_params, DESIGN_SHEETS  # noqa: E402


def _split(tmp_path):
    return split_params(PARAMS, tmp_path / "design.xlsx", tmp_path / "scenario.xlsx")


def test_sheet_partition(tmp_path):
    d_path, s_path = _split(tmp_path)
    d = openpyxl.load_workbook(d_path, read_only=True)
    s = openpyxl.load_workbook(s_path, read_only=True)
    assert set(d.sheetnames) == {"cell_params", "sections", "panel", "topology"}
    assert set(s.sheetnames) == {
        "losses", "radiation_fluxes", "mission_orbit", "mission_param",
        "analysis", "requirement", "document", "structure",
        "layer_state", "layer_shade", "layer_life", "layer_incidence",
    }


def _rows(path, sheet):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    return [tuple(r) for r in wb[sheet].iter_rows(values_only=True)]


def test_values_copied_verbatim(tmp_path):
    d_path, s_path = _split(tmp_path)
    assert _rows(d_path, "cell_params") == _rows(PARAMS, "cell_params")
    assert _rows(s_path, "losses") == _rows(PARAMS, "losses")
    assert _rows(d_path, "topology") == _rows(PARAMS, "panel")  # panel duplicated


def test_refuses_to_clobber(tmp_path):
    _split(tmp_path)
    with pytest.raises(FileExistsError):
        _split(tmp_path)
    # but overwrite=True succeeds
    split_params(PARAMS, tmp_path / "design.xlsx", tmp_path / "scenario.xlsx",
                 overwrite=True)


def test_design_sheets_routing_constant():
    assert "cell_params" in DESIGN_SHEETS and "sections" in DESIGN_SHEETS
    assert "losses" not in DESIGN_SHEETS
