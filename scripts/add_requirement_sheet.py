# -*- coding: utf-8 -*-
"""Add the `requirement` sheet to params.xlsx (idempotent).

Mirrors the authoritative workbook screenshot: a key-value sheet
(``param | type | value``) stating the mission targets the analysis checks the
array against -- operating voltage, EOR/EOL minimum power, max section current,
magnetic-moment limit, sun angle and flux at the array.

Usage:
    python scripts/add_requirement_sheet.py [path/to/params.xlsx]
"""
import sys
from pathlib import Path

import openpyxl

_HEADER = ["param", "type", "value"]
_ROWS = [
    ["voltage_operating",        "float",  101.5],
    ["voltage_unit",             "string", "V"],
    ["max_section_current",      "float",  5.4],
    ["max_section_current_unit", "string", "A"],
    ["magnetic_moment_max",      "float",  1],
    ["magnetic_moment_unit",     "string", "Am^2"],
    ["eor_power_min",            "float",  8350],
    ["eol_power_min",            "float",  7550],
    ["power_unit",               "string", "W"],
    ["sun_angle",                "float",  23.5],
    ["sun_angle_unit",           "string", "deg"],
    ["flux_at_array",            "float",  1321],
    ["flux_unit",                "string", "W/m^2"],
]


def main() -> int:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else None
    if path is None:
        root = Path(__file__).resolve().parent.parent
        for cand in (root / "params.xlsx", root / "examples" / "params.xlsx"):
            if cand.is_file():
                path = cand
                break
    if path is None or not path.is_file():
        sys.exit("ERROR: params.xlsx not found; pass its path explicitly.")

    wb = openpyxl.load_workbook(path)
    if "requirement" in wb.sheetnames:
        print("requirement: already present (left unchanged)")
    else:
        ws = wb.create_sheet("requirement")
        ws.append(_HEADER)
        for r in _ROWS:
            ws.append(r)
        print("requirement: created with %d rows" % len(_ROWS))
    wb.save(path)
    print("saved -> %s" % path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
