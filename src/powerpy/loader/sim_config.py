"""The 'panel' settings sheet in params.xlsx: panel topology + global knobs.

One key-value sheet defines the whole panel ONCE so both setup_sim.py (which
sizes the condition layers) and write_results.py (which analyses) read the same
numbers and can never get out of sync:

    param         value
    n_blocks      1
    n_parallel    4
    n_series      10
    irradiance    1.0      (global sun level; 1.0 = full / AM0)
    imp_sigma     0.0      (manufacturing Imp spread; 0 = off)
    pmax_sigma    0.0      (manufacturing Pmax spread; 0 = off)
    variance_seed 0
"""
from __future__ import annotations

from pathlib import Path
from typing import Union

import openpyxl

SHEET = "panel"

# field -> (type, default)
_FIELDS = {
    "n_blocks":      ("int", 1),
    "n_parallel":    ("int", 1),
    "n_series":      ("int", 1),
    "irradiance":    ("float", 1.0),
    "imp_sigma":     ("float", 0.0),
    "pmax_sigma":    ("float", 0.0),
    "variance_seed": ("int", 0),
}


def read_panel_config(params_path: Union[Path, str]) -> dict:
    """Read the 'panel' sheet -> dict of typed values (missing keys -> defaults).

    Raises ``ValueError`` if the workbook has no 'panel' sheet.
    """
    wb = openpyxl.load_workbook(str(params_path), data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        raise ValueError(
            "params workbook has no '%s' sheet; create it first "
            "(ensure_panel_sheet or setup_sim.py --init)" % SHEET)
    raw = {}
    for row in wb[SHEET].iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        if key in _FIELDS:
            raw[key] = row[1]
    out = {}
    for k, (typ, default) in _FIELDS.items():
        v = raw.get(k, default)
        if v is None:
            v = default
        out[k] = int(v) if typ == "int" else float(v)
    return out


SHEET_TOPOLOGY = "topology"


def read_topology(design_path: Union[Path, str]) -> dict:
    """Read the 'topology' sheet (falling back to legacy 'panel') -> typed dict.

    Same key-value format and keys as :func:`read_panel_config`, plus one
    optional ``grid_file`` key: a grid layout JSON path (absolute, or relative
    to the workbook's folder) for spatial arrays. Returns the 7 panel keys plus
    ``"grid_file"`` (``str | None``). Raises ``ValueError`` when neither sheet
    exists.
    """
    wb = openpyxl.load_workbook(str(design_path), data_only=True, read_only=True)
    if SHEET_TOPOLOGY in wb.sheetnames:
        sheet = SHEET_TOPOLOGY
    elif SHEET in wb.sheetnames:
        sheet = SHEET
    else:
        raise ValueError(
            "workbook has no '%s' or legacy '%s' sheet; create one first "
            "(split_params.py copies 'panel' as 'topology')" % (SHEET_TOPOLOGY, SHEET))
    raw = {}
    for row in wb[sheet].iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        if key in _FIELDS or key == "grid_file":
            raw[key] = row[1]
    out = {}
    for k, (typ, default) in _FIELDS.items():
        v = raw.get(k, default)
        if v is None:
            v = default
        out[k] = int(v) if typ == "int" else float(v)
    gf = raw.get("grid_file")
    out["grid_file"] = str(gf).strip() if gf not in (None, "") else None
    return out


def resolve_layout(cfg: dict, *, base_dir: Union[Path, str]):
    """Turn a :func:`read_topology` dict into a PanelLayout.

    ``grid_file`` set -> load that grid layout JSON (relative paths resolve
    against ``base_dir``, the workbook's folder); otherwise build the uniform
    ``n_blocks x n_parallel x n_series`` panel.
    """
    from powerpy.config.layout import load_layout, panel_from_topology
    gf = cfg.get("grid_file")
    if gf:
        p = Path(gf)
        if not p.is_absolute():
            p = Path(base_dir) / p
        return load_layout(str(p))
    return panel_from_topology(n_blocks=cfg["n_blocks"],
                               n_parallel=cfg["n_parallel"],
                               n_series=cfg["n_series"])


def ensure_panel_sheet(params_path: Union[Path, str], **overrides) -> bool:
    """Add a 'panel' sheet (with defaults / overrides) if missing. Returns True if created.

    Existing sheets and values are preserved; an existing 'panel' sheet is left
    untouched (so user edits are never clobbered).
    """
    path = Path(params_path)
    wb = openpyxl.load_workbook(str(path))
    if SHEET in wb.sheetnames:
        return False
    ws = wb.create_sheet(SHEET)
    ws.append(["param", "value"])
    for k, (typ, default) in _FIELDS.items():
        ws.append([k, overrides.get(k, default)])
    wb.save(str(path))
    return True
